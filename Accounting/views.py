from MBP.views import ProtectedModelViewSet
from .models import Account, Transaction
from .serializers import AccountSerializer, TransactionSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum, F,  ExpressionWrapper, DecimalField
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType
from datetime import timedelta
import calendar
from rest_framework import status, permissions
# --- Import Models from other Apps ---
from Billing.models import Invoice
from staff.models import Payroll
from inventory.models import PurchaseOrderItem
from Hotel.models import Booking
from Restaurant.models import RestaurantOrder
from Events.models import Event

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io
from decimal import Decimal

class AccountViewSet(ProtectedModelViewSet):
    queryset = Account.objects.all()
    serializer_class = AccountSerializer
    model_name = 'Account'
    lookup_field = 'slug'


class TransactionViewSet(ProtectedModelViewSet):
    queryset = Transaction.objects.select_related('account').all()
    serializer_class = TransactionSerializer
    model_name = 'Transaction'
    lookup_field = 'slug'


class DashboardStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_month_range(self, year, month):
        start_date = timezone.datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = timezone.datetime(year, month, last_day, 23, 59, 59)
        return timezone.make_aware(start_date), timezone.make_aware(end_date)

    def calculate_metrics(self, year, month, user):
        start_date, end_date = self.get_month_range(year, month)

        # --- TERA LOGIC YAHA HAI (FILTERING) ---
        hotel_obj = None
        is_superuser = False

        if user.is_superuser:
            is_superuser = True
        
        elif hasattr(user, 'hotel') and user.hotel:
            hotel_obj = user.hotel
        
        else:
            # Agar na Superuser hai, na Hotel Owner, to sab 0 return kar do
            # Note: Return values are 0 (int) here, which is fine as this is the exit path.
            return {
                "room_revenue": 0, "restaurant_revenue": 0,
                "staff_costs": 0, "operating_expenses": 0,
                "total_revenue": 0, "total_expenses": 0, "net_profit": 0
            }

        # --- 1. PREPARE QUERY FILTERS ---
        booking_filter = {}
        order_filter = {}
        event_filter = {}
        payroll_filter = {}
        inventory_filter = {}
        account_filter = {}

        if not is_superuser:
            booking_filter = {'hotel': hotel_obj}
            order_filter = {'hotel': hotel_obj}
            event_filter = {'created_by': user} 
            payroll_filter = {'staff__hotel': hotel_obj}
            inventory_filter = {'admin': user}
            account_filter = {'account__admin': user}

        # --- 2. REVENUE CALCULATION ---
        
        try:
            # Make sure these ContentType imports are working fine
            from django.contrib.contenttypes.models import ContentType
            from Hotel.models import Booking
            from Restaurant.models import RestaurantOrder
            from Events.models import Event
            booking_type = ContentType.objects.get_for_model(Booking)
            order_type = ContentType.objects.get_for_model(RestaurantOrder)
            event_type = ContentType.objects.get_for_model(Event)
        except Exception:
            # Agar koi model missing hai toh calculation skip kar do
            booking_type = None; order_type = None; event_type = None

        # A. Room Revenue
        room_revenue = Decimal('0')
        if booking_type:
            # Step 1: Filter Bookings (All or Hotel Specific)
            # ✅ Zaroori FIX: .values_list('pk', flat=True) se sirf UUIDs aayengi.
            target_booking_ids = Booking.objects.filter(**booking_filter).values_list('pk', flat=True)
            room_revenue = Invoice.objects.filter(
            content_type=booking_type,
            # object_id__in mein sirf IDs honi chahiye
            object_id__in=target_booking_ids, 
            issued_at__range=[start_date, end_date]
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        # B. Restaurant Revenue
        restaurant_revenue = Decimal('0')
        if order_type:
            target_orders = RestaurantOrder.objects.filter(**order_filter)
            restaurant_revenue = Invoice.objects.filter(
                content_type=order_type,
                object_id__in=target_orders,
                issued_at__range=[start_date, end_date]
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        # C. Event Revenue
        event_revenue = Decimal('0')
        if event_type:
            target_events = Event.objects.filter(**event_filter)
            event_revenue = Invoice.objects.filter(
                content_type=event_type,
                object_id__in=target_events,
                issued_at__range=[start_date, end_date]
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        total_revenue = room_revenue + restaurant_revenue + event_revenue

        # --- 3. EXPENSE CALCULATION ---

        # A. Staff Costs
        staff_costs = Payroll.objects.filter(
            month=month, 
            year=year, 
            **payroll_filter
        ).aggregate(total=Sum('total_salary'))['total'] or Decimal('0')

        # B. Inventory Costs 
        inventory_costs = PurchaseOrderItem.objects.filter(
            order__created_at__range=[start_date, end_date],
            **inventory_filter
        ).annotate(
            # ExpressionWrapper se database ko force karte hain ki result Decimal hi ho
            item_total=ExpressionWrapper(
                F('quantity') * F('cost_per_unit'), 
                output_field=DecimalField(max_digits=12, decimal_places=2) 
            )
        ).aggregate(total=Sum('item_total'))['total'] or Decimal('0')

        # C. General Expenses 
        general_expenses = Transaction.objects.filter(
            date__range=[start_date, end_date],
            type='debit',
            account__type='expense',
            **account_filter
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        operating_expenses = inventory_costs + general_expenses

        # --- 4. RETURN RESULT ---
        # Ab final calculations mein koi float nahi hai, sab Decimal hai.
        total_expenses = staff_costs + operating_expenses
        net_profit = total_revenue - total_expenses

        return {
            "room_revenue": round(room_revenue, 2),
            "restaurant_revenue": round(restaurant_revenue, 2),
            "staff_costs": round(staff_costs, 2),
            "operating_expenses": round(operating_expenses, 2),
            "total_revenue": round(total_revenue, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2)
        }

    def get(self, request):
        user = request.user
        now = timezone.now()
        
        # Calculate Current Stats
        current_stats = self.calculate_metrics(now.year, now.month, user)
        
        # Calculate KPIs
        rev = current_stats['total_revenue']
        exp = current_stats['total_expenses']
        prof = current_stats['net_profit']
        
        profit_margin = (prof / rev * 100) if rev > 0 else 0
        roi = (prof / exp * 100) if exp > 0 else 0

        # Calculate History
        history_reports = []
        for i in range(6):
            prev_date = now - timedelta(days=i*30)
            y, m = prev_date.year, prev_date.month
            stats = self.calculate_metrics(y, m, user)
            
            history_reports.append({
                "month": calendar.month_name[m],
                "year": y,
                "revenue": stats['total_revenue'],
                "expenses": stats['total_expenses'],
                "profit": stats['net_profit'],
                "status": "Ongoing" if (y == now.year and m == now.month) else "Completed"
            })

        return Response({
            "financial_overview": {
                "room_revenue": current_stats['room_revenue'],
                "restaurant_revenue": current_stats['restaurant_revenue'],
                "operating_expenses": current_stats['operating_expenses'],
                "staff_costs": current_stats['staff_costs'],
            },
            "quick_stats": {
                "net_profit_mtd": current_stats['net_profit'],
                "profit_margin": round(profit_margin, 1),
                "roi": round(roi, 1)
            },
            "monthly_reports": history_reports
        }, status=status.HTTP_200_OK)
    


# EXPORT TO EXCEL VIEW
class ExportExcelView(DashboardStatsView):
    """
    Inherits logic from DashboardStatsView to reuse 'calculate_metrics'
    and generates an .xlsx file.
    """
    def get(self, request):
        user = request.user
        now = timezone.now()

        # 1. Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # 2. Add Headers
        headers = ["Month", "Year", "Revenue", "Expenses", "Net Profit", "Status"]
        ws.append(headers)

        # Style Headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Blue color
            cell.alignment = Alignment(horizontal="center")

        # 3. Loop Last 6 Months (Same logic as Dashboard)
        for i in range(6):
            prev_date = now - timedelta(days=i*30)
            y, m = prev_date.year, prev_date.month
            
            # 🔥 REUSING LOGIC from Parent Class
            stats = self.calculate_metrics(y, m, user)
            
            status_text = "Ongoing" if (y == now.year and m == now.month) else "Completed"
            
            # Append Row
            ws.append([
                calendar.month_name[m],
                y,
                stats['total_revenue'],
                stats['total_expenses'],
                stats['net_profit'],
                status_text
            ])

        # 4. Prepare Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Financial_Report_{now.date()}.xlsx'
        
        wb.save(response)
        return response



# DOWNLOAD PDF VIEW
class DownloadPDFView(DashboardStatsView):
    """
    Inherits logic from DashboardStatsView to reuse 'calculate_metrics'
    and generates a .pdf file.
    """
    def get(self, request):
        user = request.user
        now = timezone.now()
        
        # 1. Setup PDF Buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # 2. Title
        styles = getSampleStyleSheet()
        title = Paragraph(f"Monthly Financial Report - {now.date()}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))

        # 3. Prepare Data for Table
        data = [["Month", "Year", "Revenue", "Expenses", "Net Profit", "Status"]] # Header

        # Loop Last 6 Months
        for i in range(6):
            prev_date = now - timedelta(days=i*30)
            y, m = prev_date.year, prev_date.month
            
            # 🔥 REUSING LOGIC
            stats = self.calculate_metrics(y, m, user)
            status_text = "Ongoing" if (y == now.year and m == now.month) else "Completed"

            data.append([
                calendar.month_name[m],
                str(y),
                f"{stats['total_revenue']}",
                f"{stats['total_expenses']}",
                f"{stats['net_profit']}",
                status_text
            ])

        # 4. Create Table
        table = Table(data, colWidths=[80, 50, 100, 100, 100, 80])
        
        # Style the Table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), # Header Background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), # Header Text
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige), # Rows Background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(style)
        
        elements.append(table)

        # 5. Build PDF
        doc.build(elements)
        
        # 6. Return Response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Financial_Report_{now.date()}.pdf'
        
        return response
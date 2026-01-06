from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Count, F, Avg, Q, DurationField, ExpressionWrapper
from django.utils import timezone
from MBP.views import ProtectedModelViewSet
from .models import Invoice, InvoiceItem, Payment
from .serializers import InvoiceSerializer, InvoiceItemSerializer, PaymentSerializer, RecentActivitySerializer
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.db.models.functions import TruncMonth
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from datetime import datetime,date, timedelta
import calendar
# Models import
from Hotel.models import Booking, Hotel,Room,RoomCategory,Guest
from Restaurant.models import RestaurantOrder,OrderItem
from Reviews.models import RestaurantReview, HotelReview
from MBP.models import AuditLog
from MBP.serializers import AuditLogSerializer
from django.utils.timesince import timesince

class InvoiceViewSet(ProtectedModelViewSet):
    serializer_class = InvoiceSerializer
    model_name = 'Invoice'
    lookup_field = 'slug'
    queryset = Invoice.objects.all()
    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()

        # 1️⃣ Superuser → all invoices
        if user.is_superuser:
            return qs

        role = getattr(user, "role", None)
        if not role:
            return Invoice.objects.none()

        role_name = role.name.lower()

        # 2️⃣ Hotel Admin (Owner)
        # Admin jis hotel ka owner hai,
        # us hotel ke staff / users ke invoices
        if role_name == "admin":
            return qs.filter(
                issued_to__hotel=user.hotel
            ).distinct()

        # 3️⃣ Staff
        # Staff apne hotel ke invoices dekhe
        if role_name == "staff":
            if hasattr(user, "staff_profile") and user.staff_profile.hotel:
                return qs.filter(
                    # Q (issued_to__hotel=user.staff_profile.hotel) |
                    Q(created_by=user)
                ).distinct()
            return Invoice.objects.none()

        # 4️⃣ Customer
        # Customer sirf apne invoices dekhe
        if role_name == "customer":
            return qs.filter(
                issued_to=user
            )

        # 5️⃣ Others (vendor, unknown roles)
        return Invoice.objects.none()
    

    @action(detail=True, methods=['patch'], url_path='pay')
    def pay_invoice(self, request, slug=None):
        invoice = self.get_object()
        amount = request.data.get('amount_paid', None)

        if amount is None:
            return Response({"error": "amount_paid is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            amount = float(amount)
        except ValueError:
            return Response({"error": "Invalid amount format"}, status=status.HTTP_400_BAD_REQUEST)

        if amount <= 0:
            return Response({"error": "Amount must be greater than zero"}, status=status.HTTP_400_BAD_REQUEST)

        #  Update amount paid
        invoice.amount_paid += amount
        invoice.save()

        serializer = self.get_serializer(invoice)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='summary')
    def invoice_summary(self, request):
        user = request.user
        
        # Content type to EXCLUDE
        exclude_ct = ContentType.objects.get(app_label='inventory', model='purchaseorder')

        if not user.is_staff and not user.is_superuser:
            qs = Invoice.objects.filter(issued_to=user).exclude(content_type=exclude_ct)

        # USER IS ADMIN  only invoices created for his own hotel/business
        elif user.is_staff and not user.is_superuser:
            qs = Invoice.objects.filter(created_by=user).exclude(content_type=exclude_ct)

        # USER IS SUPERUSER all invoices from all admins (except Purchase Order)
        else:
            qs = Invoice.objects.exclude(content_type=exclude_ct)
       
        # TOTAL INVOICES (issued_to user)
        total_invoices = qs.count()

        # PENDING
        pending_invoices = qs.filter(status='unpaid').count()

        # OVERDUE
        overdue_invoices = qs.filter(
            status='unpaid',
            due_date__lt=timezone.now().date()
        ).count()

        # ⭐ TOTAL REVENUE = sum of paid invoice amounts EXCEPT inventory purchase order invoices
        total_revenue = qs.filter(
            status='paid'
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        return Response({
            "total_revenue": float(total_revenue),
            "total_invoices": total_invoices,
            "pending_invoices": pending_invoices,
            "overdue_invoices": overdue_invoices
        })
    
    @action(detail=False, methods=['post'], url_path='export')
    def export_invoices(self, request):
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

        if not start_date or not end_date:
            return Response({"error": "start_date and end_date are required."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Filter invoices
        invoices = Invoice.objects.filter(
            issued_at__date__gte=start_date,
            issued_at__date__lte=end_date
        ).order_by('-issued_at')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        # Header row
        headers = [
            "Invoice No", "Customer Name", "Issued Date",
            "Due Date", "Total Amount", "Amount Paid","Balance Due", "Status"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Insert invoice data
        for row, invoice in enumerate(invoices, start=2):
            balance_due = float(invoice.total_amount) - float(invoice.amount_paid)

            ws.cell(row=row, column=1, value=invoice.slug)
            ws.cell(row=row, column=2, value=invoice.customer_name)
            ws.cell(row=row, column=3, value=str(invoice.issued_at.date()))
            ws.cell(row=row, column=4, value=str(invoice.due_date))
            ws.cell(row=row, column=5, value=float(invoice.total_amount))
            ws.cell(row=row, column=6, value=float(invoice.amount_paid))
            ws.cell(row=row, column=7, value=balance_due)
            ws.cell(row=row, column=8, value=invoice.status)

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="invoices_{start_date}_to_{end_date}.xlsx"'

        wb.save(response)
        return response
    


    @action(detail=False, methods=['get'], url_path='recent-invoices')
    def recent_invoices(self, request):
        today = timezone.now().date()
        
        invoices = Invoice.objects.filter(
            issued_at__date=today
        ).order_by('-issued_at')

        serializer = self.get_serializer(invoices, many=True)
        return Response(serializer.data)



class InvoiceItemViewSet(ProtectedModelViewSet):
    queryset = InvoiceItem.objects.all().select_related('invoice')
    serializer_class = InvoiceItemSerializer
    model_name = 'InvoiceItem'
    lookup_field = 'slug'


class PaymentViewSet(ProtectedModelViewSet):
    queryset = Payment.objects.all().select_related('invoice')
    serializer_class = PaymentSerializer
    model_name = 'Payment'
    lookup_field = 'slug'

    def get_queryset(self):
        user = self.request.user
        qs = Payment.objects.select_related('invoice')

        # 1️⃣ Superuser → all payments
        if user.is_superuser:
            return qs

        # 2️⃣ Admin / Staff → only their hotel payments
        hotel = None

        # Admin (hotel owner)
        if hasattr(user, 'role') and user.role and user.role.name.lower() == 'admin':
            hotel = getattr(user, 'hotel', None)

        # Staff (assigned hotel)
        elif hasattr(user, 'staff_profile') and getattr(user.staff_profile, 'hotel', None):
            hotel = user.staff_profile.hotel

        if hotel:
            return qs.filter(
                Q(invoice__issued_to__hotel=hotel) |
                Q(invoice__related_object__hotel=hotel)
            ).distinct()

        # 3️⃣ Others (customer, vendor, etc.) → no access
        return qs.none()
    
    @action(detail=False, methods=['get'], url_path='today-revenue')
    def today_revenue(self, request):
        today = timezone.now().date()

        revenue = Payment.objects.filter(payment_date__date=today).aggregate(
            total_revenue=Sum('amount_paid')
        )['total_revenue'] or 0

        return Response({
            "date": today,
            "total_revenue": round(revenue, 2)
        })


class RevenueAnalyticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        # Default to current year if not provided
        year = request.query_params.get('year', datetime.now().year)

        target_hotel = None
        hotel_name_display = "All Hotels (Global Revenue)"

        # --- 1. Identify User Scope ---
        if user.is_superuser:
            # Case A: Superuser wants specific hotel data
            hotel_id = request.query_params.get('hotel_id')
            if hotel_id:
                target_hotel = Hotel.objects.filter(id=hotel_id).first()
                if target_hotel:
                    hotel_name_display = target_hotel.name
            # Case B: Superuser didn't send hotel_id -> target_hotel remains None (Global View)
            
        else:
            # Case C: Regular Admin/Staff (Restricted to their hotel)
            if hasattr(user, 'role') and user.role.name.lower() == 'admin':
                target_hotel = getattr(user, 'hotel', None)
            elif hasattr(user, 'staff_profile') and user.staff_profile.hotel:
                target_hotel = user.staff_profile.hotel
            
            if not target_hotel:
                return Response({"error": "No hotel associated with this user."}, status=403)
            
            hotel_name_display = target_hotel.name

        # --- 2. Filter IDs (Booking & Orders) ---
        # Hum pehle IDs nikalenge taaki Invoice table me GenericForeignKey se match kar sakein
        
        booking_qs = Booking.objects.filter(created_at__year=year)
        order_qs = RestaurantOrder.objects.filter(order_time__year=year)

        # Agar specific hotel select hua hai (Admin ya Superuser with filter), to filter lagao
        if target_hotel:
            booking_qs = booking_qs.filter(hotel=target_hotel)
            order_qs = order_qs.filter(hotel=target_hotel)
        
        # IDs list nikalo
        booking_ids = booking_qs.values_list('id', flat=True)
        order_ids = order_qs.values_list('id', flat=True)

        # --- 3. Content Types Fetch ---
        booking_ct = ContentType.objects.get_for_model(Booking)
        order_ct = ContentType.objects.get_for_model(RestaurantOrder)

        # --- 4. Aggregation Logic (Group by Month) ---
        
        # Hotel Revenue
        hotel_revenue_data = Invoice.objects.filter(
            content_type=booking_ct,
            object_id__in=booking_ids,
            status='paid',  # Only count paid invoices
            issued_at__year=year
        ).annotate(month=TruncMonth('issued_at')).values('month').annotate(total=Sum('total_amount')).order_by('month')

        # Restaurant Revenue
        restaurant_revenue_data = Invoice.objects.filter(
            content_type=order_ct,
            object_id__in=order_ids,
            status='paid',
            issued_at__year=year
        ).annotate(month=TruncMonth('issued_at')).values('month').annotate(total=Sum('total_amount')).order_by('month')

        # --- 5. Format Data for Graph (Jan-Dec) ---
        final_data = []
        
        # Convert DB Query to Dictionary {MonthNum: Amount}
        h_map = {item['month'].month: item['total'] for item in hotel_revenue_data}
        r_map = {item['month'].month: item['total'] for item in restaurant_revenue_data}

        months = list(calendar.month_abbr)[1:]  # ['Jan', 'Feb', ... 'Dec']

        for index, month_name in enumerate(months, start=1):
            final_data.append({
                "month": month_name,
                "hotel_revenue": h_map.get(index, 0),
                "restaurant_revenue": r_map.get(index, 0)
            })

        return Response({
            "year": year,
            "hotel_name": hotel_name_display,
            "currency": "₹",
            "chart_data": final_data
        })
    
class TopAnalyticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        year = request.query_params.get('year', datetime.now().year)

        target_hotel = None
        
        # --- 1. Identify User Scope ---
        if user.is_superuser:
            hotel_id = request.query_params.get('hotel_id')
            if hotel_id:
                target_hotel = Hotel.objects.filter(id=hotel_id).first()
        else:
            if hasattr(user, 'role') and user.role.name.lower() == 'admin':
                target_hotel = getattr(user, 'hotel', None)
            elif hasattr(user, 'staff_profile') and user.staff_profile.hotel:
                target_hotel = user.staff_profile.hotel
            
            if not target_hotel:
                return Response({"error": "No hotel associated with this user."}, status=403)

        # =================================================
        # SECTION A: Top Menu Items (Restaurant)
        # =================================================
        
        menu_filters = {
            'order__status__in': ['served', 'completed'], 
            'order__order_time__year': year
        }

        if target_hotel:
            menu_filters['order__hotel'] = target_hotel

        # 1. Get Top items by Revenue
        menu_items_qs = OrderItem.objects.filter(**menu_filters).values(
            name=F('menu_item__name'),
            item_id=F('menu_item__id') # ✅ Item ID bhi le rahe hain rating fetch karne ke liye
        ).annotate(
            total_orders=Count('order__id', distinct=True), 
            revenue=Sum(F('price') * F('quantity'))
        ).order_by('-revenue')[:5]

        # 2. Format Menu Data with ACTUAL Rating
        top_menu_data = []
        for item in menu_items_qs:
            # ✅ REAL RATING CALCULATION
            # RestaurantReview table check karega us specific menu item ke liye
            rating_data = RestaurantReview.objects.filter(
                menu_item__id=item['item_id']
            ).aggregate(avg=Avg('rating'))
            
            # Agar rating hai to round off karo, nahi to 0
            actual_rating = round(rating_data['avg'], 1) if rating_data['avg'] else 0

            top_menu_data.append({
                "name": item['name'],
                "orders": item['total_orders'],
                "rating": actual_rating, # 👈 Ab ye DB se aayega
                "revenue": item['revenue'] or 0
            })

        # =================================================
        # SECTION B: Top Performing Rooms (Hotel)
        # =================================================
        
        if target_hotel:
            categories = RoomCategory.objects.filter(hotel=target_hotel)
        else:
            categories = RoomCategory.objects.all()

        booking_ct = ContentType.objects.get_for_model(Booking)
        top_rooms_data = []

        start_date = date(int(year), 1, 1)
        if int(year) == datetime.now().year:
            end_date = datetime.now().date()
        else:
            end_date = date(int(year), 12, 31)
        
        total_days = (end_date - start_date).days + 1

        for cat in categories:
            booking_filters = {
                'room__room_category': cat,
                'check_in__year': year,
                'status__in': ['checked_in', 'checked_out', 'confirmed']
            }
            
            cat_bookings = Booking.objects.filter(**booking_filters)

            cat_revenue = Invoice.objects.filter(
                content_type=booking_ct,
                object_id__in=cat_bookings.values('id'),
                status='paid'
            ).aggregate(sum=Sum('total_amount'))['sum'] or 0

            total_rooms_in_cat = cat.rooms.count()
            
            if total_rooms_in_cat > 0:
                total_nights_sold = sum(bk.total_nights for bk in cat_bookings)
                max_capacity_nights = total_rooms_in_cat * total_days
                occupancy_rate = (total_nights_sold / max_capacity_nights) * 100
                occupancy_rate = round(min(occupancy_rate, 100), 1)
            else:
                occupancy_rate = 0

            if cat_revenue > 0 or cat_bookings.exists():
                top_rooms_data.append({
                    "name": cat.name,
                    "hotel_name": cat.hotel.name if not target_hotel else "",
                    "bookings": cat_bookings.count(),
                    "occupancy": f"{occupancy_rate}%",
                    "revenue": cat_revenue
                })

        top_rooms_data.sort(key=lambda x: x['revenue'], reverse=True)
        top_rooms_data = top_rooms_data[:5]

        return Response({
            "scope": "Global" if not target_hotel else target_hotel.name,
            "menu_items": top_menu_data,
            "rooms": top_rooms_data
        })
    
class DashboardStatsView(APIView):
    """
    API for Top Dashboard Cards:
    1. Monthly Revenue (Total) + Growth %
    2. Average Occupancy % + Growth %
    3. Restaurant Revenue + Growth %
    4. Customer Satisfaction (Avg Rating) + Growth (Absolute Change)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        target_hotel = None
        
        # --- 1. Identify User Scope ---
        if user.is_superuser:
            hotel_id = request.query_params.get('hotel_id')
            if hotel_id:
                target_hotel = Hotel.objects.filter(id=hotel_id).first()
        else:
            if hasattr(user, 'role') and user.role.name.lower() == 'admin':
                target_hotel = getattr(user, 'hotel', None)
            elif hasattr(user, 'staff_profile') and user.staff_profile.hotel:
                target_hotel = user.staff_profile.hotel
            
            if not target_hotel:
                return Response({"error": "No hotel associated with this user."}, status=403)

        # --- 2. Date Ranges (Current Month vs Last Month) ---
        now = timezone.now()
        
        # Current Month
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month_start = (current_month_start + timedelta(days=32)).replace(day=1)
        
        # Last Month
        last_month_end = current_month_start - timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # --- Helper: Calculate Growth ---
        def calculate_growth(current, previous):
            if previous == 0:
                return "+100%" if current > 0 else "0%"
            change = ((current - previous) / previous) * 100
            sign = "+" if change >= 0 else ""
            return f"{sign}{round(change, 1)}%"

        # --- Helper: Get Revenue ---
        def get_revenue(start_date, end_date, hotel=None, content_types=None):
            # Base filters
            filters = {
                'issued_at__gte': start_date,
                'issued_at__lt': end_date,
                'status': 'paid' # Only realized revenue
            }
            
            # Agar specific hotel hai, toh humein content_type se filter karna padega
            if hotel:
                # 1. Booking IDs
                booking_ids = Booking.objects.filter(hotel=hotel).values_list('id', flat=True)
                # 2. Order IDs
                order_ids = RestaurantOrder.objects.filter(hotel=hotel).values_list('id', flat=True)
                
                # Complex query: Invoice linked to Booking OR Invoice linked to Order
                # Note: ContentType filter zaroori hai taaki mix na ho
                booking_ct = ContentType.objects.get_for_model(Booking)
                order_ct = ContentType.objects.get_for_model(RestaurantOrder)
                
                q_obj = Q(content_type=booking_ct, object_id__in=booking_ids) | \
                        Q(content_type=order_ct, object_id__in=order_ids)
                
                return Invoice.objects.filter(q_obj, **filters).aggregate(sum=Sum('total_amount'))['sum'] or 0
            
            # Global View (Superuser without filter)
            return Invoice.objects.filter(**filters).aggregate(sum=Sum('total_amount'))['sum'] or 0

        # --- Helper: Get Specific Restaurant Revenue ---
        def get_rest_revenue(start_date, end_date, hotel=None):
            order_ct = ContentType.objects.get_for_model(RestaurantOrder)
            filters = {
                'issued_at__gte': start_date,
                'issued_at__lt': end_date,
                'status': 'paid',
                'content_type': order_ct
            }
            
            if hotel:
                order_ids = RestaurantOrder.objects.filter(hotel=hotel).values_list('id', flat=True)
                filters['object_id__in'] = order_ids
            
            return Invoice.objects.filter(**filters).aggregate(sum=Sum('total_amount'))['sum'] or 0

        # --- Helper: Get Occupancy % ---
        def get_occupancy(start_date, end_date, hotel=None):
            # Total Rooms
            room_qs = Room.objects.all()
            if hotel:
                room_qs = room_qs.filter(hotel=hotel)
            total_rooms = room_qs.count()
            
            if total_rooms == 0: return 0

            # Days in range
            days = (end_date - start_date).days
            total_capacity = total_rooms * days
            
            # Total Booked Nights
            booking_qs = Booking.objects.filter(
                check_in__lt=end_date,
                check_out__gt=start_date,
                status__in=['checked_in', 'checked_out', 'confirmed']
            )
            if hotel:
                booking_qs = booking_qs.filter(hotel=hotel)
            
            total_booked_nights = 0
            # Rough calculation for range overlap
            # Production me isse better logic lagana pad sakta hai agar heavy traffic ho
            for bk in booking_qs:
                # Intersection of booking dates and current month
                overlap_start = max(bk.check_in, start_date.date())
                overlap_end = min(bk.check_out, end_date.date())
                nights = (overlap_end - overlap_start).days
                if nights > 0:
                    total_booked_nights += nights
            
            return (total_booked_nights / total_capacity) * 100

        # --- Helper: Get Satisfaction Rating ---
        def get_rating(start_date, end_date, hotel=None):
            # Hotel Reviews
            h_qs = HotelReview.objects.filter(date__gte=start_date, date__lt=end_date)
            # Restaurant Reviews
            r_qs = RestaurantReview.objects.filter(date__gte=start_date, date__lt=end_date)
            
            if hotel:
                h_qs = h_qs.filter(hotel=hotel)
                r_qs = r_qs.filter(menu_item__category__hotel=hotel)

            h_avg = h_qs.aggregate(avg=Avg('rating'))['avg']
            r_avg = r_qs.aggregate(avg=Avg('rating'))['avg']
            
            # Simple Average of both categories
            values = [v for v in [h_avg, r_avg] if v is not None]
            if not values:
                return 0
            return sum(values) / len(values)


        # ==========================================
        # 1. MONTHLY REVENUE (Total)
        # ==========================================
        curr_rev = get_revenue(current_month_start, next_month_start, target_hotel)
        prev_rev = get_revenue(last_month_start, current_month_start, target_hotel)
        rev_growth = calculate_growth(curr_rev, prev_rev)

        # ==========================================
        # 2. RESTAURANT REVENUE
        # ==========================================
        curr_rest_rev = get_rest_revenue(current_month_start, next_month_start, target_hotel)
        prev_rest_rev = get_rest_revenue(last_month_start, current_month_start, target_hotel)
        rest_growth = calculate_growth(curr_rest_rev, prev_rest_rev)

        # ==========================================
        # 3. AVERAGE OCCUPANCY
        # ==========================================
        curr_occ = get_occupancy(current_month_start, next_month_start, target_hotel)
        prev_occ = get_occupancy(last_month_start, current_month_start, target_hotel)
        
        # Absolute diff for occupancy usually looks better (e.g. +5%)
        occ_diff = curr_occ - prev_occ
        occ_sign = "+" if occ_diff >= 0 else ""
        occ_growth = f"{occ_sign}{round(occ_diff, 1)}%"

        # ==========================================
        # 4. CUSTOMER SATISFACTION
        # ==========================================
        curr_rating = get_rating(current_month_start, next_month_start, target_hotel)
        prev_rating = get_rating(last_month_start, current_month_start, target_hotel)
        
        rating_diff = curr_rating - prev_rating
        rating_sign = "+" if rating_diff >= 0 else ""
        rating_growth = f"{rating_sign}{round(rating_diff, 1)}"


        # ==========================================
        # FINAL RESPONSE
        # ==========================================
        return Response({
            "monthly_revenue": {
                "value": curr_rev,
                "growth": rev_growth
            },
            "average_occupancy": {
                "value": f"{round(curr_occ)}%",
                "growth": occ_growth
            },
            "restaurant_revenue": {
                "value": curr_rest_rev,
                "growth": rest_growth
            },
            "customer_satisfaction": {
                "value": f"{round(curr_rating, 1)}/5",
                "growth": rating_growth
            }
        })

class CustomerAnalyticsView(APIView):
    """
    API for Customer Analytics Dashboard Cards.
    Logic: Based on Check-in Date and Guest Email History.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        target_hotel = None

        # ==========================================
        # 1. ISOLATION LOGIC (Admin vs Superuser)
        # ==========================================
        if user.is_superuser:
            hotel_id = request.query_params.get('hotel_id')
            if hotel_id:
                target_hotel = Hotel.objects.filter(id=hotel_id).first()
        else:
            if hasattr(user, 'role') and user.role.name.lower() == 'admin':
                target_hotel = getattr(user, 'hotel', None)
            elif hasattr(user, 'staff_profile') and user.staff_profile.hotel:
                target_hotel = user.staff_profile.hotel
            
            if not target_hotel:
                return Response({"error": "No hotel associated with this user."}, status=403)

        # ==========================================
        # 2. DATE RANGES (Based on Check-in)
        # ==========================================
        now = timezone.now()
        
        # Current Month Range
        curr_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if curr_start.month == 12:
            next_month_start = curr_start.replace(year=curr_start.year + 1, month=1)
        else:
            next_month_start = curr_start.replace(month=curr_start.month + 1)
        
        # Last Month Range
        last_month_end = curr_start - timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # ==========================================
        # 3. CORE LOGIC FUNCTION
        # ==========================================
        def get_metrics_for_range(start_date, end_date):
            """
            Calculates metrics for bookings having check-in between start_date and end_date.
            """
            
            # --- A. Filter Bookings by CHECK-IN DATE ---
            # Hum sirf Valid bookings count karenge (Checked In, Checked Out, Confirmed)
            qs = Booking.objects.filter(
                check_in__gte=start_date, 
                check_in__lt=end_date,
                status__in=['checked_in', 'checked_out', 'confirmed']
            ).select_related('hotel').prefetch_related('guests')
            
            if target_hotel:
                qs = qs.filter(hotel=target_hotel)

            new_count = 0
            ret_count = 0

            # --- B. New vs Returning Logic (Loop through Bookings) ---
            for booking in qs:
                # 1. Get First Guest Email
                first_guest = booking.guests.order_by('created_at').first()
                
                if not first_guest or not first_guest.email:
                    # Agar guest details hi nahi h, ya email nahi h -> Default to New
                    new_count += 1
                    continue
                
                email = first_guest.email.lower().strip()
                current_check_in = booking.check_in

                # 2. Check History in GUEST Table
                # Query: Kya is email wala koi guest, is booking ki date se PEHLE aaya h?
                history_qs = Guest.objects.filter(
                    email=email,
                    booking__check_in__lt=current_check_in, # Strictly BEFORE this visit
                    booking__status__in=['checked_in', 'checked_out', 'confirmed'] # Only count valid past stays
                )

                # Hotel Isolation for History
                if target_hotel:
                    history_qs = history_qs.filter(booking__hotel=target_hotel)

                # 3. Decision
                if history_qs.exists():
                    ret_count += 1
                else:
                    new_count += 1

            # --- C. Average Stay Logic ---
            # Avg Stay = (Checkout - Checkin) for bookings in this range
            avg_stay_data = qs.annotate(
                duration=ExpressionWrapper(
                    F('check_out') - F('check_in'),
                    output_field=DurationField()
                )
            ).aggregate(avg=Avg('duration'))
            
            avg_stay_days = 0.0
            if avg_stay_data['avg']:
                avg_stay_days = round(avg_stay_data['avg'].total_seconds() / 86400, 1)

            # --- D. CLV / Avg Invoice Value ---
            # In bookings ke invoices ka average
            booking_ct = ContentType.objects.get_for_model(Booking)
            booking_ids = qs.values_list('id', flat=True)
            
            avg_inv_data = Invoice.objects.filter(
                content_type=booking_ct,
                object_id__in=booking_ids,
                status='paid' # Sirf paid invoices ka average lenge
            ).aggregate(avg_val=Avg('total_amount'))
            
            avg_invoice_val = float(avg_inv_data['avg_val']) if avg_inv_data['avg_val'] else 0.0

            return {
                "new": new_count,
                "returning": ret_count,
                "avg_stay": avg_stay_days,
                "clv": avg_invoice_val
            }

        # ==========================================
        # 4. EXECUTE CALCULATIONS
        # ==========================================
        curr_data = get_metrics_for_range(curr_start, next_month_start)
        prev_data = get_metrics_for_range(last_month_start, curr_start)

        # Helper for Growth Strings
        def calc_growth(current, previous, is_percentage=True):
            if previous == 0:
                return "+100%" if current > 0 else "0%"
            
            diff = current - previous
            
            if is_percentage:
                percent = (diff / previous) * 100
                sign = "+" if percent >= 0 else ""
                return f"{sign}{round(percent, 1)}%"
            else:
                # Absolute change for Avg Stay (e.g. +0.4)
                sign = "+" if diff >= 0 else ""
                return f"{sign}{round(diff, 1)}"

        # ==========================================
        # 5. FINAL RESPONSE
        # ==========================================
        return Response({
            "new_customers": {
                "value": curr_data['new'],
                "growth": calc_growth(curr_data['new'], prev_data['new'])
            },
            "returning_customers": {
                "value": curr_data['returning'],
                "retention_rate": calc_growth(curr_data['returning'], prev_data['returning']),
            },
            "average_stay": {
                "value": f"{curr_data['avg_stay']} days",
                "growth": calc_growth(curr_data['avg_stay'], prev_data['avg_stay'], is_percentage=False)
            },
            "customer_lifetime_value": {
                "value": curr_data['clv'],
                "growth": calc_growth(curr_data['clv'], prev_data['clv'])
            }
        })
    
class RecentActivityViewSet(ProtectedModelViewSet):
    """
    Dashboard logic using ProtectedModelViewSet.
    Base model is AuditLog because the activity feed is based on logs.
    """
    # Base ordering is important
    queryset = AuditLog.objects.all().order_by('-timestamp')
    serializer_class = RecentActivitySerializer
    model_name = 'AuditLog'

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        
        if not user.is_authenticated:
            return qs.none()

        if user.is_superuser:
            return qs
            
        created_users = getattr(user, 'created_users', None)
        if created_users:
             qs = qs.filter(Q(user=user) | Q(user__in=created_users.all()))
        else:
             qs = qs.filter(user=user)
        return qs

    # @action(detail=False, methods=['get'], url_path='recent-activity')
    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        relevant_models = ['Booking', 'HotelReview', 'Invoice']
        qs = qs.filter(model_name__in=relevant_models)

        # Time Logic
        # one_hour_ago = timezone.now() - timedelta(minutes=60)
        
        # Check for logs in last 60 mins
        # recent_logs_qs = qs.filter(timestamp__gte=one_hour_ago)
        
        final_data = qs[:5]

        # if recent_logs_qs.exists():
        #     #  Agar recent logs hain, tab bhi max 10 hi dikhana (Safety)
        #     final_data = recent_logs_qs[:5]
        # # else:
        # #     #  Agar recent nahi hain, to last 4 records uthao (Strict Slice)
        # #     final_data = qs[:4]

        serializer = self.get_serializer(final_data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
        
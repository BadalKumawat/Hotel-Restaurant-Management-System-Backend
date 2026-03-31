from MBP.views import ProtectedModelViewSet
from .models import Lead, Customer, Interaction
from .serializers import LeadSerializer, CustomerSerializer, InteractionSerializer
from MBP.models import AuditLog
from rest_framework.response import Response
from django.utils import timezone
from django.utils.timesince import timesince
from rest_framework.decorators import action
from Reviews.models import HotelReview, RestaurantReview, ServiceReview
from datetime import timedelta
from django.db.models import Avg,Count,Q
import csv
import io # Text stream handle karne ke liye
from openpyxl import load_workbook # Excel ke liye (Lightweight than Pandas)
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.db import transaction
from Hotel.models import Guest, Booking
from rest_framework.exceptions import ValidationError




class LeadViewSet(ProtectedModelViewSet):
    queryset = Lead.objects.all()
    serializer_class = LeadSerializer
    model_name = 'Lead'
    lookup_field = 'slug'

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()

        if not user.is_authenticated:
            return qs.none()

        if user.is_superuser:
            pass

        # Admin / Staff → own hotel
        elif hasattr(user, "role") and user.role and user.role.name.lower() in ["admin"]:
            if hasattr(user, 'hotel') and user.hotel:
                qs = qs.filter(hotel=user.hotel)
            else:
                return qs.none()
            
        elif hasattr(user, "role") and user.role and user.role.name.lower() == "staff":
            if hasattr(user, "staff_profile"):
                return qs.filter(hotel=user.staff_profile.hotel)

        # Vendor / Customer → NO access
        else:
            return qs.none()

    # 2. AUTO ASSIGN HOTEL
    def perform_create(self, serializer):
        user = self.request.user
        hotel = getattr(user, 'hotel', None)
        serializer.save(created_by=user, hotel=hotel)


class CustomerViewSet(ProtectedModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    model_name = 'Customer'
    lookup_field = 'slug'

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()

        if not user.is_authenticated:
            return qs.none()

        if user.is_superuser:
            pass

        # Admin / Staff → own hotel
        elif hasattr(user, "role") and user.role and user.role.name.lower() in ["admin"]:
            if hasattr(user, 'hotel') and user.hotel:
                qs = qs.filter(hotel=user.hotel)
            else:
                return qs.none()
            
        elif hasattr(user, "role") and user.role and user.role.name.lower() == "staff":
            if hasattr(user, "staff_profile"):
                return qs.filter(hotel=user.staff_profile.hotel)

        # Vendor / Customer → NO access
        else:
            return qs.none()
        
        # B. Search (Name, Email, Phone)
        search_query = self.request.query_params.get('search', None)
        if search_query:
            qs = qs.filter(
                Q(name__icontains=search_query) | 
                Q(email__icontains=search_query) | 
                Q(phone__icontains=search_query)
            )

        # C. Status Filter (Active, Inactive, VIP, Regular)
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            val = status_filter.lower()
            if val == 'vip':
                qs = qs.filter(customer_type='vip')
            elif val == 'regular':
                qs = qs.filter(customer_type='regular')
            elif val == 'active':
                qs = qs.filter(status='active')
            elif val == 'inactive':
                qs = qs.filter(status='inactive')

        return qs
    
    # ------------------------------------------------
    
    # SAVE KARTE WAQT HOTEL ASSIGN KARNA
    def perform_create(self, serializer):
        user = self.request.user
        hotel = getattr(user, 'hotel', None)
        serializer.save(created_by=user, hotel=hotel)

    def perform_destroy(self, instance):
        # Check Unpaid Invoices
        # Assuming Invoice is linked to User Email
        from Billing.models import Invoice
        has_unpaid = Invoice.objects.filter(
            issued_to__email=instance.email,
            status__in=['unpaid', 'partial']
        ).exists()
        
        if has_unpaid:
            raise ValidationError("Cannot delete customer with unpaid invoices.")
        
        instance.delete()

    @action(detail=False, methods=['get'], url_path='recent-activities')
    def recent_activities(self, request):
        """
        Dashboard ke liye Last 10 Customer Activities fetch karta hai.
        """
        
        user = request.user
        user_hotel = getattr(user, 'hotel', None)
        # 1. Define karein ki kaunse Models ki activity dikhani hai
        relevant_models = ['Booking', 'RestaurantOrder', 'Invoice', 'Payment', 'Review', 'Guest']
        
        # 2. Database se logs fetch karein (Latest first)
        # Hum user__isnull=False check kar rahe hain taaki system logs na aayein
        logs_qs = AuditLog.objects.filter(
            model_name__in=relevant_models,
            user__isnull=False 
        )

        if not user.is_superuser and user_hotel:
            logs_qs = logs_qs.filter(user__hotel=user_hotel)

        logs = logs_qs.select_related('user').order_by('-timestamp')[:10]

        activities = []

        for log in logs:
            # 3. Message format karna (Technical -> Human Readable)
            action_text = ""
            icon_type = "activity" # UI ke liye icon hint (optional)

            # --- Logic to generate message ---
            if log.model_name == 'Booking':
                icon_type = "calendar"
                if log.action == 'create':
                    action_text = "Made a room reservation"
                elif log.action == 'update':
                    if 'checked_in' in (log.details or ""):
                        action_text = "Checked in to Hotel"
                    elif 'checked_out' in (log.details or ""):
                        action_text = "Checked out"
                    else:
                        action_text = "Updated booking details"
            
            elif log.model_name == 'RestaurantOrder':
                icon_type = "food"
                if log.action == 'create':
                    action_text = "Ordered food at Restaurant"
                elif 'completed' in (log.details or ""):
                    action_text = "Completed dining"
            
            elif log.model_name == 'Invoice':
                icon_type = "bill"
                if log.action == 'create':
                    action_text = "Invoice generated"
                elif 'paid' in (log.details or ""):
                    action_text = "Paid the invoice"

            elif log.model_name == 'Review':
                icon_type = "star"
                action_text = "Left a review"

            else:
                # Fallback for other actions
                action_text = f"{log.action.title()}d {log.model_name}"

            # 4. List mein append karna
            activities.append({
                "id": log.id,
                "customer_name": log.user.full_name if log.user else "Guest",
                "customer_avatar":  None, # Agar user.profile.image ho to yahan bhej sakte hain
                "action_text": action_text,
                "time_ago": timesince(log.timestamp, timezone.now()) + " ago",
                "icon": icon_type,
                "model": log.model_name
            })

        return Response(activities)
    

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        # Base Queryset (Already filtered by get_queryset for Hotel)
        qs = self.filter_queryset(self.get_queryset())
        
        now = timezone.now()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_end = start_of_month - timedelta(days=1)
        start_of_last_month = last_month_end.replace(day=1)

        # 1. TOTAL CUSTOMERS
        total_customers = qs.count()
        last_month_total = qs.filter(created_at__lte=last_month_end).count()
        
        total_growth = 0
        if last_month_total > 0:
            total_growth = ((total_customers - last_month_total) / last_month_total) * 100

        # 2. VIP CUSTOMERS
        vip_count = qs.filter(customer_type='vip').count()
        last_month_vip = qs.filter(
            customer_type='vip', created_at__lte=last_month_end
        ).count()

        vip_growth = 0
        if last_month_vip > 0:
            vip_growth = ((vip_count - last_month_vip) / last_month_vip) * 100

        # 3. ACTIVE THIS MONTH
        active_count = qs.filter(last_visit__gte=start_of_month).count()
        last_month_active = qs.filter(
            last_visit__gte=start_of_last_month, last_visit__lte=last_month_end
        ).count()

        active_growth = 0
        if last_month_active > 0:
            active_growth = ((active_count - last_month_active) / last_month_active) * 100

        # 4. CUSTOMER SATISFACTION (Filtered by Hotel)
        user_hotel = getattr(request.user, 'hotel', None)

        def get_weighted_rating(cutoff_date=None, hotel=None):
            if not hotel: return 0.0

            filters = {'hotel': hotel} # Only reviews for THIS hotel
            if cutoff_date:
                filters['date__lte'] = cutoff_date
            
            # Hotel Reviews
            h_stats = HotelReview.objects.filter(**filters).aggregate(avg=Avg('rating'), count=Count('id'))
            h_avg = h_stats['avg'] or 0
            h_count = h_stats['count'] or 0

            # Restaurant Reviews (Assuming items belong to hotel categories)
            # Complex filter: RestaurantReview -> MenuItem -> MenuCategory -> Hotel
            r_stats = RestaurantReview.objects.filter(menu_item__category__hotel=hotel).aggregate(avg=Avg('rating'), count=Count('id'))
            if cutoff_date:
                 r_stats = RestaurantReview.objects.filter(menu_item__category__hotel=hotel, date__lte=cutoff_date).aggregate(avg=Avg('rating'), count=Count('id'))

            r_avg = r_stats['avg'] or 0
            r_count = r_stats['count'] or 0

            # Service Reviews (Generic service, might not be directly linked to hotel easily without logic, skipping for simplicity or assuming linked via User)
            # For now relying on Hotel Reviews mainly + Restaurant
            
            total_reviews = h_count + r_count
            if total_reviews == 0:
                return 0.0
            
            total_stars = (h_avg * h_count) + (r_avg * r_count)
            return total_stars / total_reviews

        current_rating = get_weighted_rating(hotel=user_hotel)
        last_month_rating = get_weighted_rating(cutoff_date=last_month_end, hotel=user_hotel)

        satisfaction_change = round(current_rating - last_month_rating, 1)

        data = {
            "total_customers": {
                "value": total_customers,
                "trend": round(total_growth, 1),
                "trend_direction": "up" if total_growth >= 0 else "down"
            },
            "vip_customers": {
                "value": vip_count,
                "trend": round(vip_growth, 1),
                "trend_direction": "up" if vip_growth >= 0 else "down"
            },
            "active_monthly": {
                "value": active_count,
                "trend": round(active_growth, 1),
                "trend_direction": "up" if active_growth >= 0 else "down"
            },
            "satisfaction": {
                "value": round(current_rating, 1),
                "trend": satisfaction_change,
                "trend_direction": "up" if satisfaction_change >= 0 else "down"
            }
        }
        return Response(data)

    # ---------------------------------------------------
    # 5. EXPORT API (Filtered by Hotel)
    # ---------------------------------------------------
    @action(detail=False, methods=['get'], url_path='export')
    def export_data(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers List"

        headers = ['Name', 'Email', 'Phone', 'Address', 'City', 'Country', 'Type', 'Preferences']
        ws.append(headers)

        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Uses get_queryset() so it is already filtered by hotel
        customers = self.filter_queryset(self.get_queryset())
        
        for customer in customers.iterator():
            ws.append([
                customer.name,
                customer.email,
                customer.phone,
                customer.address,
                customer.city,
                customer.country,
                customer.customer_type,
                customer.preferences
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customers_list.xlsx"'
        return response

    # ---------------------------------------------------
    # 6. IMPORT API (Assigns to Admin's Hotel)
    # ---------------------------------------------------
    @action(detail=False, methods=['post'], url_path='import')
    def import_data(self, request):
        file_obj = request.FILES.get('file')
        user = request.user
        user_hotel = getattr(user, 'hotel', None)

        if not file_obj:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
        
        # VALIDATION File Size (Max 5MB)
        if file_obj.size > 5 * 1024 * 1024:
             return Response({"error": "File too large. Max size is 5MB."}, status=status.HTTP_400_BAD_REQUEST)
        
        if not user_hotel and not user.is_superuser:
            return Response({"error": "You must be associated with a hotel to import customers."}, status=status.HTTP_403_FORBIDDEN)

        created_count = 0
        skipped_count = 0
        errors = []
        rows_data = []

        try:
            if file_obj.name.endswith('.csv'):
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                # VALIDATION 2: Row Limit (Max 1000)
                if len(decoded_file) > 1001: 
                     return Response({"error": "Too many rows. Max 1000 customers allowed per import."}, status=status.HTTP_400_BAD_REQUEST)
                rows_data = list(csv.DictReader(decoded_file))


            elif file_obj.name.endswith(('.xlsx', '.xls')):
                wb = load_workbook(file_obj, data_only=True)
                ws = wb.active
                # VALIDATION 2: Row Limit
                if ws.max_row > 1001:
                    return Response({"error": "Too many rows. Max 1000 customers allowed per import."}, status=status.HTTP_400_BAD_REQUEST)
                
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows_data.append({headers[i]: value for i, value in enumerate(row) if i < len(headers)})
            
            else:
                return Response({"error": "Invalid format. Use .csv or .xlsx"}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                for index, row in enumerate(rows_data):
                    try:
                        email = str(row.get('Email', '') or '').strip()
                        name = str(row.get('Name', '') or '').strip()
                        phone = str(row.get('Phone', '') or '').strip()

                        if not email or not name: continue

                        # Phone Validation inside loop
                        import re
                        if phone and not re.match(r'^\+?1?\d{9,15}$', phone):
                            errors.append(f"Row {index+2}: Invalid Phone {phone}")
                            continue

                        # Check Duplicate WITHIN THIS HOTEL
                        if Customer.objects.filter(email=email, hotel=user_hotel).exists():
                            skipped_count += 1
                            continue

                        Customer.objects.create(
                            hotel=user_hotel, # Assign to current admin's hotel
                            name=name,
                            email=email,
                            phone=phone[:20],
                            address=str(row.get('Address', '') or ''),
                            city=str(row.get('City', '') or ''),
                            country=str(row.get('Country', '') or ''),
                            customer_type=str(row.get('Type', 'regular') or 'regular').lower(),
                            status='active',
                            preferences=str(row.get('Preferences', '') or '')
                        )
                        created_count += 1
                        
                    except Exception as inner_e:
                        errors.append(f"Row {index + 2}: {str(inner_e)}")

            return Response({
                "status": "success",
                "created": created_count,
                "skipped": skipped_count,
                "errors": errors[:5]
            })

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ---------------------------------------------------
    # 7. SYNC DATABASE (Syncs only Hotel's Guests)
    # ---------------------------------------------------
    @action(detail=False, methods=['post'], url_path='sync-database')
    def sync_database(self, request):
        user = request.user
        user_hotel = getattr(user, 'hotel', None)

        if not user_hotel and not user.is_superuser:
             return Response({"error": "You must be associated with a hotel to sync data."}, status=status.HTTP_403_FORBIDDEN)

        created_count = 0
        skipped_count = 0
        
        # 1. Booking Users (Filtered by Hotel)
        booking_users = Booking.objects.filter(hotel=user_hotel).values('user__email', 'user__full_name', 'user__phone').distinct()
        
        for user_data in booking_users:
            email = user_data.get('user__email')
            name = user_data.get('user__full_name') or "Unknown User"
            phone = user_data.get('user__phone') or ""

            if email and not Customer.objects.filter(email=email, hotel=user_hotel).exists():
                Customer.objects.create(
                    hotel=user_hotel, # Assign Hotel
                    name=name,
                    email=email,
                    phone=phone,
                    customer_type='regular'
                )
                created_count += 1
            else:
                skipped_count += 1

        # 2. Guests (Filtered by Hotel)
        # Guest -> Booking -> Hotel
        guests = Guest.objects.filter(booking__hotel=user_hotel).values('email', 'first_name', 'last_name', 'phone', 'address', 'id_proof_type').distinct()
        
        for guest in guests:
            email = guest.get('email')
            if not email: continue 
                
            full_name = f"{guest.get('first_name')} {guest.get('last_name') or ''}".strip()
            phone = guest.get('phone') or ""
            address = guest.get('address') or ""
            
            if not Customer.objects.filter(email=email, hotel=user_hotel).exists():
                Customer.objects.create(
                    hotel=user_hotel, # Assign Hotel
                    name=full_name,
                    email=email,
                    phone=phone,
                    address=address,
                    customer_type='regular',
                    preferences=f"ID Proof: {guest.get('id_proof_type') or 'N/A'}"
                )
                created_count += 1
            else:
                skipped_count += 1

        return Response({
            "status": "Sync Completed",
            "new_customers_created": created_count,
            "already_existed": skipped_count,
        })


class InteractionViewSet(ProtectedModelViewSet):
    queryset = Interaction.objects.all()
    serializer_class = InteractionSerializer
    model_name = 'Interaction'
    lookup_field = 'slug'

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()

        if not user.is_authenticated:
            return qs.none()

        if user.is_superuser:
            pass

        # Admin / Staff → own hotel
        elif hasattr(user, "role") and user.role and user.role.name.lower() in ["admin"]:
            if hasattr(user, 'hotel') and user.hotel:
                qs = qs.filter(hotel=user.hotel)
            else:
                return qs.none()
            
        elif hasattr(user, "role") and user.role and user.role.name.lower() == "staff":
            if hasattr(user, "staff_profile"):
                return qs.filter(hotel=user.staff_profile.hotel)

        # Vendor / Customer → NO access
        else:
            return qs.none()

    # 2. AUTO ASSIGN HOTEL
    def perform_create(self, serializer):
        user = self.request.user
        hotel = getattr(user, 'hotel', None)
        
        # Ensure referenced customer belongs to same hotel
        customer_slug = self.request.data.get('customer')
        if customer_slug:
             # Serializer validate karega, but extra safety
             pass
             
        serializer.save(handled_by=user, hotel=hotel)
